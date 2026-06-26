"""Write a SuiteReport back into the suite sheet as a timestamped RESULTS block.

Each call appends a new block below any existing content:

    RESULTS — run <YYYY-MM-DD HH:MM:SS>
    test_id | status | actual_status | expected_status | detail
    <one row per case>

The parser stops reading cases at the first ``RESULTS`` marker (see ``tools/suite.py``), so
stacked blocks never interfere with re-parsing the suite.

Safety: a timestamped backup is taken first, and after saving the case-definition region is
re-read and compared to its pre-write state. If the save altered it (e.g. the known
numbers-parser smart-quote round-trip glitch), the file is restored from the backup and a
:class:`ResultsWriteError` is raised — the sheet is never left corrupted.
"""

from __future__ import annotations

import datetime
import json
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl.styles import Font, PatternFill

from ..models import CaseEvidence, CaseReport, SuiteReport
from .suite import RESULTS_MARKER, _as_str, _find_header_row, _load_rows

RESULTS_HEADER = [
    "test_id", "status", "actual_status", "expected_status", "correlation_id", "detail",
]

# ✅/❌ status icons render in colour everywhere; the fills/fonts below add the rest of the polish.
PASS_ICON = "✅ PASS"
FAIL_ICON = "❌ FAIL"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


# Cell styles (font, optional fill) keyed by role — mirrors the worker's xlsx-js-style palette.
_STYLES: dict[str, tuple[Font, PatternFill | None]] = {
    "banner": (Font(bold=True, size=13, color="FFFFFFFF"), _fill("FF1F3864")),
    "header": (Font(bold=True, color="FF1F3864"), _fill("FFD9E1F2")),
    "pass": (Font(bold=True, color="FF0B6E2D"), _fill("FFE6F4EA")),
    "fail": (Font(bold=True, color="FFB3261E"), _fill("FFFCE8E6")),
    "section": (Font(bold=True, color="FF1F3864"), _fill("FFEFEFEF")),
    "label": (Font(bold=True, color="FF555555"), None),
    "title": (Font(bold=True, size=12, color="FF1F3864"), None),
}


def _style(cell, role: str) -> None:
    font, fill = _STYLES[role]
    cell.font = font
    if fill is not None:
        cell.fill = fill


class ResultsWriteError(Exception):
    """Writing the results block failed or would have corrupted the sheet."""


def write_results(path: str, report: SuiteReport, run_at: str | None = None) -> str:
    """Append a timestamped results block to the suite at ``path``. Returns the run timestamp."""
    run_at = run_at or datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    block = _build_block(report, run_at)

    backup = file_path.with_name(
        f"{file_path.stem}.bak-{datetime.datetime.now():%Y%m%d-%H%M%S}{file_path.suffix}"
    )
    shutil.copy2(file_path, backup)

    before = _case_region(file_path)
    if suffix == ".numbers":
        _append_numbers(file_path, block)
    elif suffix in {".xlsx", ".xlsm"}:
        _append_xlsx(file_path, block)
    else:
        backup.unlink(missing_ok=True)
        raise ResultsWriteError(f"unsupported file type for write-back: {suffix}")

    after = _case_region(file_path)
    if after != before:
        shutil.copy2(backup, file_path)  # restore the untouched original
        raise ResultsWriteError(
            f"save altered the test-definition rows; restored from backup ({backup.name})"
        )

    backup.unlink(missing_ok=True)  # verified intact — the backup is no longer needed
    return run_at


# --- block construction ----------------------------------------------------------------


def _build_block(report: SuiteReport, run_at: str) -> list[list[str]]:
    # Keep the cell starting with "RESULTS" (the parser's stop marker) — the icon goes at the end.
    overall = "✅" if report.failed == 0 else "❌"
    summary = f"RESULTS — run {run_at}  (passed {report.passed}/{report.total}) {overall}"
    rows: list[list[str]] = [[summary], list(RESULTS_HEADER)]
    for case in report.cases:
        rows.append([
            case.test_id,
            PASS_ICON if case.passed else FAIL_ICON,
            "" if case.actual_status is None else str(case.actual_status),
            "" if case.expected_status is None else str(case.expected_status),
            case.correlation_id or "",
            _detail(case),
        ])
    return rows


def _detail(case: CaseReport) -> str:
    if case.error:
        return case.error
    parts: list[str] = []
    ra = case.response_assert
    if ra is not None and not ra.passed:
        if not ra.status_ok:
            parts.append("status mismatch")
        missing = [d.path for d in ra.diffs if d.message == "missing key"]
        if missing:
            parts.append("missing keys: " + ", ".join(missing))
        mismatched = [d.path for d in ra.diffs if d.message == "value mismatch"]
        if mismatched:
            parts.append("value mismatch: " + ", ".join(mismatched))
    lv = case.log_validation
    if lv is not None:
        if lv.missing:
            parts.append("missing logs: " + ", ".join(lv.missing))
        elif lv.used_fallback:
            parts.append("logs ok (whole-log fallback)")
        else:
            parts.append("logs ok")
    if not parts:
        mode = ra.mode if ra is not None else ""
        return f"response matched ({mode})" if case.passed else "did not match"
    return "; ".join(parts)


# --- writers ---------------------------------------------------------------------------


def _append_numbers(file_path: Path, block: list[list[str]]) -> None:
    from numbers_parser import Document

    doc = Document(str(file_path))
    sheet = next((s for s in doc.sheets if s.name.lower() == "tests"), doc.sheets[0])
    table = sheet.tables[0]
    start = table.num_rows
    for _ in range(1 + len(block)):  # +1 leaves a blank separator row
        table.add_row()
    for i, line in enumerate(block):
        for c, val in enumerate(line):
            table.write(start + 1 + i, c, val)
    doc.save(str(file_path))


def _append_xlsx(file_path: Path, block: list[list[str]]) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_path)
    sheet = next(
        (workbook[name] for name in workbook.sheetnames if name.lower() == "tests"),
        workbook.active,
    )
    # ``sheet.append`` anchors on openpyxl's tracked max row, which for sheets exported from
    # Google Sheets is padded with ~1000 phantom empty rows — that would bury the block far
    # below the data. Anchor on the true last non-empty row and write from there instead.
    row = _last_data_row(sheet) + 2  # +1 blank separator, then the block
    for line in block:
        cells = []
        for col, val in enumerate(line, start=1):
            cells.append(sheet.cell(row=row, column=col, value=val))
        _style_block_row(sheet, row, line, cells)
        row += 1
    _widen(sheet, {"A": 16, "B": 12, "C": 13, "D": 14, "E": 22, "F": 64})
    workbook.save(file_path)


def _style_block_row(sheet, row: int, line: list[str], cells: list) -> None:
    """Colour a RESULTS-block row: banner, header, or the PASS/FAIL status cell."""
    first = line[0] if line else ""
    if first.lower().startswith(RESULTS_MARKER):
        _style(cells[0], "banner")
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=len(RESULTS_HEADER)
        )
    elif line == RESULTS_HEADER:
        for cell in cells:
            _style(cell, "header")
    elif len(line) > 1 and line[1] in (PASS_ICON, FAIL_ICON):
        _style(cells[1], "pass" if line[1] == PASS_ICON else "fail")


def _widen(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _last_data_row(sheet) -> int:
    """The 1-based index of the last row holding any non-empty cell (0 if the sheet is empty)."""
    last = 0
    for r in range(1, sheet.max_row + 1):
        if any(
            (sheet.cell(row=r, column=c).value not in (None, ""))
            for c in range(1, sheet.max_column + 1)
        ):
            last = r
    return last


# --- verification ----------------------------------------------------------------------


def _case_region(file_path: Path) -> list[list[str]]:
    """The rows from the top through the last test-case row (before any RESULTS marker).

    Comparing this region before/after the write detects any save-time corruption of the
    test definitions.
    """
    rows = _load_rows(file_path)
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return [[_as_str(c) or "" for c in r] for r in rows]

    end = len(rows)
    for i in range(header_idx + 1, len(rows)):
        first = _as_str(rows[i][0]) if rows[i] else None
        if first and first.lower().startswith(RESULTS_MARKER):
            end = i
            break

    region = [_normalize_row(r) for r in rows[:end]]
    # Drop trailing blank rows so an added separator row before the results block doesn't
    # register as a change to the test definitions.
    while region and not region[-1]:
        region.pop()
    return region


# --- per-case evidence tabs ------------------------------------------------------------


def write_evidence_tabs(path: str, evidence: list[CaseEvidence], run_at: str) -> None:
    """Overwrite one sheet tab per test case with the latest run's evidence.

    Each case gets its own tab (named after its ``test_id``) holding the request, the response
    validation, and the actual log lines that matched — overwritten in place so only the latest
    run is kept. This is additive to (and independent of) the stacked ``RESULTS`` summary block
    that ``write_results`` appends to the ``tests`` sheet.

    xlsx/.xlsm only; a no-op for other formats. The ``tests`` definition region is backed up and
    verified unchanged (it lives on a different sheet, so this is a belt-and-braces guard).
    """
    file_path = Path(path)
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"} or not evidence:
        return

    backup = file_path.with_name(
        f"{file_path.stem}.bak-{datetime.datetime.now():%Y%m%d-%H%M%S}-evi{file_path.suffix}"
    )
    shutil.copy2(file_path, backup)

    before = _case_region(file_path)
    _write_evidence_xlsx(file_path, evidence, run_at)
    after = _case_region(file_path)
    if after != before:
        shutil.copy2(backup, file_path)
        raise ResultsWriteError(
            f"evidence write altered the test-definition rows; restored ({backup.name})"
        )

    backup.unlink(missing_ok=True)


def _write_evidence_xlsx(file_path: Path, evidence: list[CaseEvidence], run_at: str) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_path)
    used: set[str] = set()
    for ev in evidence:
        name = _safe_sheet_name(ev.test_id, used)
        if name in workbook.sheetnames:
            workbook.remove(workbook[name])  # override: keep only the latest evidence
        sheet = workbook.create_sheet(title=name)
        _fill_evidence_sheet(sheet, ev, run_at)
    workbook.save(file_path)


_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(test_id: str, used: set[str]) -> str:
    """An Excel-legal, stable, unique tab name for a test id (<=31 chars, no ``[]:*?/\\``)."""
    base = (_INVALID_SHEET_CHARS.sub("_", test_id).strip() or "case")[:31]
    if base.casefold() == "tests":  # never collide with the suite sheet
        base = f"{base}_evi"[:31]
    name, n = base, 2
    while name.casefold() in used:
        suffix = f"~{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name.casefold())
    return name


def _fill_evidence_sheet(sheet, ev: CaseEvidence, run_at: str) -> None:
    """Lay out one case's evidence as vertical key/value sections."""
    result_icon = PASS_ICON if ev.passed else FAIL_ICON
    rows: list[list[Any]] = [
        [f"{ev.test_id} — evidence", f"run {run_at}", f"RESULT: {result_icon}"],
    ]
    if ev.description:
        rows.append([ev.description])
    if ev.error:
        rows.append(["error", ev.error])

    rows += [[], ["📋 [Request]"]]
    rows.append(["method", ev.method or ""])
    rows.append(["url", ev.url or ""])
    rows.append(["headers", _json(ev.request_headers)])
    rows.append(["body", _json(ev.request_body)])

    if ev.response_passed is None:
        resp_status = ""
    else:
        resp_status = PASS_ICON if ev.response_passed else FAIL_ICON
    rows += [[], ["🔎 [Response validation]", resp_status]]
    rows.append(["expected_status", _s(ev.expected_status)])
    rows.append(["actual_status", _s(ev.actual_status)])
    rows.append(["match_mode", _s(ev.match_mode)])
    rows.append(["latency_ms", "" if ev.latency_ms is None else _s(round(ev.latency_ms, 1))])
    # Expected vs actual response bodies side by side, so a case can be validated at a glance.
    rows += [[], ["expected_result", _json(ev.expected_response)]]
    rows.append(["actual_result", _json(ev.actual_body)])
    if ev.response_diffs:
        rows.append(["diffs"])
        for d in ev.response_diffs:
            rows.append(["", f"{d.path}: {d.message} "
                             f"(expected={d.expected!r}, actual={d.actual!r})"])
    else:
        rows.append(["diffs", "(none)"])

    if not ev.validated_logs:
        log_status = "not validated"
    else:
        log_status = PASS_ICON if ev.logs_passed else FAIL_ICON
    rows += [[], ["📜 [Log validation]", log_status]]
    if ev.validated_logs:
        rows.append(["log_source", ev.log_source or ""])
        rows.append(["correlation_id", ev.correlation_id or ""])
        rows.append(["used_fallback", "yes (whole-log)" if ev.used_fallback else "no"])
        rows.append(["lines_considered", _s(ev.lines_considered)])
        rows += [[], ["expected_log_string", "matched_lines"]]
        for needle in ev.expected_log_strings:
            lines = ev.matched_log_lines.get(needle, [])
            rows.append([needle, lines[0] if lines else ""])  # blank when nothing matched
            for extra in lines[1:]:
                rows.append(["", extra])

    for r, line in enumerate(rows, start=1):
        cells = {}
        for c, val in enumerate(line, start=1):
            if val not in (None, ""):
                cells[c] = sheet.cell(row=r, column=c, value=val)
        _style_evidence_row(r, line, cells)
    _widen(sheet, {"A": 22, "B": 66, "C": 16})


def _style_evidence_row(row: int, line: list[Any], cells: dict) -> None:
    """Style an evidence row: title, coloured RESULT/status, section headers and labels."""
    c0 = str(line[0]) if line else ""
    c1 = str(line[1]) if len(line) > 1 else ""
    if row == 1:
        if 1 in cells:
            _style(cells[1], "title")
        result = str(line[2]) if len(line) > 2 else ""
        if 3 in cells:
            _style(cells[3], "pass" if "PASS" in result else "fail")
    elif "[" in c0 or c0 in ("expected_log_string", "expected_result", "diffs"):
        if 1 in cells:
            _style(cells[1], "section")
        if 2 in cells:
            if "PASS" in c1:
                _style(cells[2], "pass")
            elif "FAIL" in c1:
                _style(cells[2], "fail")
            else:
                _style(cells[2], "section")
    elif c0 and 1 in cells:
        _style(cells[1], "label")


def _json(value: Any) -> str:
    """Render a header/body cell: strings as-is, dict/list pretty-printed JSON."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False, indent=2)
    except (TypeError, ValueError):
        return str(value)


def _s(value: Any) -> str:
    return "" if value is None else str(value)


def _normalize_row(row: list) -> list[str]:
    """Row as trimmed strings with trailing empties removed.

    Trailing-cell trimming matters because appending a wider results block pads the existing
    (narrower) definition rows with empty cells — that is not a content change.
    """
    cells = [_canon_cell(c) for c in row]
    while cells and cells[-1] == "":
        cells.pop()
    return cells


def _canon_cell(value) -> str:
    """Stringify a cell, treating an integer-valued float as its int form.

    Sheets exported from Google Sheets store whole numbers as floats (``201.0``); openpyxl
    rewrites them as ints (``201``) on save. That round-trip is not a content change, so the
    before/after guard must not flag it.
    """
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _as_str(value) or ""
