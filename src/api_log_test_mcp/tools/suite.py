"""read_test_suite: parse a test suite (.numbers or .xlsx) into structured TestCases.

The sheet has an optional metadata block at the top (e.g. ``Basepath | <url>``) followed by a
header row (located by finding the ``test_id`` column, not by assuming row 1) and one row per
case. The column schema is isolated here (see ``COLUMNS``) so a schema change is a localized
edit. Bad rows never abort the parse; each problem becomes a ``ParseError`` and the row is
skipped.

Expected-log encoding: the ``expected_log_strings`` cell is parsed as a JSON array first; if
that fails, it falls back to splitting on newlines or ``||``.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from ..models import LogMatchMode, MatchMode, ParseError, TestCase, TestSuite

# Canonical column header -> TestCase field. Header lookup is case-insensitive/trimmed.
# Aliases let older sheets (match_mode / expected_logs) keep working.
COLUMNS = {
    "test_id": "test_id",
    "description": "description",
    "method": "method",
    "url": "url",
    "headers": "headers",
    "body": "body",
    "auth_required": "auth_required",
    "expected_status": "expected_status",
    "expected_response": "expected_response",
    "response_match_mode": "response_match_mode",
    "match_mode": "response_match_mode",  # alias
    "ignore_paths": "ignore_paths",
    "validate_logs": "validate_logs",
    "expected_log_strings": "expected_log_strings",
    "expected_logs": "expected_log_strings",  # alias
    "log_match_mode": "log_match_mode",
    "log_source": "log_source",
}

_BOOL_TRUE = {"yes", "y", "true", "1"}
_BOOL_FALSE = {"no", "n", "false", "0"}

# A row whose test_id cell starts with this marks the end of the test-case section (e.g. a
# written-back "RESULTS — run ..." block lives below the cases in the same sheet).
RESULTS_MARKER = "results"


def read_test_suite(path: str) -> TestSuite:
    """Parse the suite at ``path`` into a TestSuite (base_path + cases + parse_errors)."""
    file_path = Path(path)
    if not file_path.exists():
        return TestSuite(parse_errors=[ParseError(row=0, message=f"file not found: {path}")])

    try:
        rows = _load_rows(file_path)
    except Exception as exc:  # noqa: BLE001 - surface any reader error as a parse error
        return TestSuite(parse_errors=[ParseError(row=0, message=f"could not read sheet: {exc}")])

    header_idx = _find_header_row(rows)
    if header_idx is None:
        return TestSuite(
            parse_errors=[ParseError(row=0, message="no header row containing 'test_id' found")]
        )

    base_path = _extract_base_path(rows[:header_idx])
    header_map = _map_headers(rows[header_idx])

    suite = TestSuite(base_path=base_path)
    seen_ids: set[str] = set()
    tid_idx = header_map["test_id"]
    for offset, raw_row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if _is_blank(raw_row):
            continue
        marker = _as_str(raw_row[tid_idx]) if tid_idx < len(raw_row) else None
        if marker and marker.lower().startswith(RESULTS_MARKER):
            break  # results block (written back below the cases) — stop parsing cases
        _parse_row(raw_row, header_map, offset, seen_ids, suite)

    return suite


# --- readers ---------------------------------------------------------------------------


def _load_rows(file_path: Path) -> list[list[Any]]:
    """Load every cell as a list-of-rows, dispatching on file extension."""
    suffix = file_path.suffix.lower()
    if suffix == ".numbers":
        return _load_numbers(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(file_path)
    raise ValueError(f"unsupported file type '{suffix}' (expected .numbers or .xlsx)")


def _load_numbers(file_path: Path) -> list[list[Any]]:
    from numbers_parser import Document

    doc = Document(str(file_path))
    sheet = next((s for s in doc.sheets if s.name.lower() == "tests"), doc.sheets[0])
    table = sheet.tables[0]
    return [list(row) for row in table.rows(values_only=True)]


def _load_xlsx(file_path: Path) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = next(
        (workbook[name] for name in workbook.sheetnames if name.lower() == "tests"),
        workbook.active,
    )
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


# --- header / metadata -----------------------------------------------------------------


def _find_header_row(rows: list[list[Any]]) -> int | None:
    """Index of the first row containing a cell equal to 'test_id' (case-insensitive)."""
    for idx, row in enumerate(rows):
        for cell in row:
            if _as_str(cell) and str(cell).strip().lower() == "test_id":
                return idx
    return None


def _extract_base_path(meta_rows: list[list[Any]]) -> str | None:
    """Find a ``Basepath | <url>`` row in the metadata block above the header."""
    for row in meta_rows:
        if not row:
            continue
        label = _as_str(row[0])
        if label and label.lower() in {"basepath", "base_path", "base path", "baseurl",
                                       "base_url", "base url"}:
            for cell in row[1:]:
                value = _as_str(cell)
                if value:
                    return value
    return None


def _map_headers(header_row: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, name in enumerate(header_row):
        key = _as_str(name)
        if not key:
            continue
        canonical = COLUMNS.get(key.strip().lower())
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


# --- row parsing -----------------------------------------------------------------------


def _parse_row(
    raw_row: list[Any],
    header_map: dict[str, int],
    row_no: int,
    seen_ids: set[str],
    suite: TestSuite,
) -> None:
    """Parse one data row, appending a TestCase or ParseError(s) to ``suite``."""

    def cell(field: str) -> Any:
        idx = header_map.get(field)
        if idx is None or idx >= len(raw_row):
            return None
        return raw_row[idx]

    test_id = _as_str(cell("test_id"))
    if not test_id:
        suite.parse_errors.append(
            ParseError(row=row_no, column="test_id", message="missing test_id")
        )
        return
    if test_id in seen_ids:
        suite.parse_errors.append(
            ParseError(row=row_no, column="test_id", message=f"duplicate test_id '{test_id}'")
        )
        return

    row_errors: list[ParseError] = []

    headers = _parse_json_cell(cell("headers"), "headers", row_no, row_errors, default={})
    body = _parse_json_cell(cell("body"), "body", row_no, row_errors, default=None,
                            allow_scalar=True)
    expected_response = _parse_json_cell(
        cell("expected_response"), "expected_response", row_no, row_errors,
        default=None, allow_scalar=True,
    )

    expected_status = _as_int(cell("expected_status"), "expected_status", row_no, row_errors)
    response_match_mode = _as_enum(cell("response_match_mode"), MatchMode, MatchMode.JSON_SUBSET,
                                   "response_match_mode", row_no, row_errors)
    log_match_mode = _as_enum(cell("log_match_mode"), LogMatchMode, LogMatchMode.CONTAINS,
                              "log_match_mode", row_no, row_errors)
    auth_required = _as_bool(cell("auth_required"), "auth_required", row_no, row_errors,
                             default=True)
    validate_logs = _as_bool(cell("validate_logs"), "validate_logs", row_no, row_errors,
                             default=False)

    if row_errors:
        suite.parse_errors.extend(row_errors)
        return

    case = TestCase(
        test_id=test_id,
        description=_as_str(cell("description")),
        method=(_as_str(cell("method")) or "GET").upper(),
        url=_as_str(cell("url")) or "",
        headers=headers if isinstance(headers, dict) else {},
        body=body,
        auth_required=auth_required,
        expected_status=expected_status,
        expected_response=expected_response,
        response_match_mode=response_match_mode,
        ignore_paths=_split_list(cell("ignore_paths")),
        validate_logs=validate_logs,
        expected_log_strings=_parse_expected_logs(cell("expected_log_strings")),
        log_match_mode=log_match_mode,
        log_source=(_as_str(cell("log_source")) or "anypoint").lower(),
    )
    seen_ids.add(test_id)
    suite.cases.append(case)


# --- cell helpers ----------------------------------------------------------------------


def _is_blank(row: list[Any]) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_int(value: Any, column: str, row_no: int, errors: list[ParseError]) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return int(float(value)) if isinstance(value, str) else int(value)
    except (TypeError, ValueError):
        errors.append(ParseError(row=row_no, column=column, message=f"not an integer: {value!r}"))
        return None


def _as_bool(value: Any, column: str, row_no: int, errors: list[ParseError], *,
             default: bool) -> bool:
    if isinstance(value, bool):
        return value
    text = _as_str(value)
    if not text:
        return default
    lowered = text.lower()
    if lowered in _BOOL_TRUE:
        return True
    if lowered in _BOOL_FALSE:
        return False
    errors.append(ParseError(row=row_no, column=column, message=f"not yes/no: {text!r}"))
    return default


def _as_enum(value: Any, enum_cls, default, column: str, row_no: int, errors: list[ParseError]):
    text = _as_str(value)
    if not text:
        return default
    try:
        return enum_cls(text.lower())
    except ValueError:
        allowed = ", ".join(e.value for e in enum_cls)
        errors.append(
            ParseError(row=row_no, column=column,
                       message=f"invalid {column} {text!r}; allowed: {allowed}")
        )
        return default


def _parse_json_cell(
    value: Any, column: str, row_no: int, errors: list[ParseError],
    *, default: Any, allow_scalar: bool = False,
) -> Any:
    """Parse a cell that should contain JSON. Empty -> default."""
    if value is None:
        return default
    if not isinstance(value, str):
        # Already a native scalar from the sheet (number/bool); keep as-is when allowed.
        return value if allow_scalar else default
    text = value.strip()
    if not text:
        return default
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        if allow_scalar:
            return text
        errors.append(ParseError(row=row_no, column=column, message="invalid JSON"))
        return default


def _split_list(value: Any) -> list[str]:
    text = _as_str(value)
    if not text:
        return []
    parts = [p.strip() for chunk in text.splitlines() for p in chunk.split(",")]
    return [p for p in parts if p]


def _parse_expected_logs(value: Any) -> list[str]:
    """JSON array first; fall back to newline / '||' delimited strings."""
    text = _as_str(value)
    if not text:
        return []
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [str(item) for item in parsed]
    except json.JSONDecodeError:
        pass
    delimiter = "||" if "||" in text else "\n"
    return [p.strip() for p in text.split(delimiter) if p.strip()]
