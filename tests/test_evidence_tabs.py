"""matched_log_lines + per-case evidence tabs (overwrite-in-place, definitions preserved)."""

from openpyxl import Workbook, load_workbook

from api_log_test_mcp.cache.snapshot_store import SnapshotStore
from api_log_test_mcp.config import LogBackend, Settings
from api_log_test_mcp.models import CaseEvidence, LogMatchMode, MatchMode
from api_log_test_mcp.tools import logs as logtools
from api_log_test_mcp.tools.results_writer import _safe_sheet_name, write_evidence_tabs
from api_log_test_mcp.tools.suite import read_test_suite

# --- matched_log_lines -----------------------------------------------------------------


def _settings(log_path: str) -> Settings:
    return Settings(log_backend=LogBackend.FILE, file_log_path=log_path)


def test_matched_log_lines_returns_actual_lines(sample_log_path):
    store = SnapshotStore()
    sid = logtools.snapshot_logs(_settings(sample_log_path), store=store)

    result = logtools.matched_log_lines(
        sid, "order-001", ["Order lookup succeeded", "this never appears"], store=store
    )
    assert len(result["Order lookup succeeded"]) == 1
    assert "Order lookup succeeded" in result["Order lookup succeeded"][0]
    assert result["this never appears"] == []  # no match -> blank evidence


def test_matched_log_lines_requires_correlation_id(tmp_path):
    # The expected message appears on two lines: one carrying the correlation id, one not.
    # Strict scoping (no fallback) must record only the correlation-id line.
    log = tmp_path / "app.log"
    log.write_text(
        "INFO event:TC-9-aaa - this is start log\n"   # corr id + message -> recorded
        "INFO event:OTHER-1 - this is start log\n"     # message only, wrong corr id -> excluded
    )
    store = SnapshotStore()
    sid = logtools.snapshot_logs(_settings(str(log)), store=store)

    result = logtools.matched_log_lines(
        sid, "TC-9-aaa", ["this is start log"], correlation_fallback=False, store=store
    )
    lines = result["this is start log"]
    assert len(lines) == 1
    assert "TC-9-aaa" in lines[0]  # only the line carrying the correlation id


def test_matched_log_lines_regex_and_fallback(sample_log_path):
    store = SnapshotStore()
    sid = logtools.snapshot_logs(_settings(sample_log_path), store=store)

    # Unknown correlation id -> whole-log fallback; regex matches every "Received request" line.
    result = logtools.matched_log_lines(
        sid, "nope", ["Received request (GET|POST)"], mode=LogMatchMode.REGEX,
        correlation_fallback=True, store=store,
    )
    # 3 inbound requests in the fixture (orders, payments, health)
    assert len(result["Received request (GET|POST)"]) == 3


# --- write_evidence_tabs ---------------------------------------------------------------


def _make_suite_xlsx(path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "tests"
    ws.append(["Basepath", "https://api.test/"])
    ws.append([])
    ws.append(["test_id", "method", "url", "expected_status"])
    ws.append(["TC-001", "POST", "/orders", 201])
    ws.append(["TC-002", "POST", "/orders", 400])
    wb.save(path)


def _evidence(status: int, matched: dict[str, list[str]]) -> CaseEvidence:
    return CaseEvidence(
        test_id="TC-001",
        passed=True,
        method="POST",
        url="https://api.test/orders",
        request_body={"sku": "ABC-100"},
        actual_status=status,
        expected_status=201,
        match_mode=MatchMode.JSON_SUBSET,
        response_passed=True,
        actual_body={"status": "ACCEPTED"},
        validated_logs=True,
        logs_passed=bool(matched),
        log_source="anypoint",
        correlation_id="TC-001-abc",
        expected_log_strings=list(matched.keys()),
        matched_log_lines=matched,
    )


def _tab_text(path, name: str) -> str:
    ws = load_workbook(path)[name]
    return "\n".join(
        "|".join("" if c is None else str(c) for c in row)
        for row in ws.iter_rows(values_only=True)
    )


def test_evidence_tab_created_with_request_and_logs(tmp_path):
    path = tmp_path / "suite.xlsx"
    _make_suite_xlsx(path)
    ev = _evidence(201, {"start log": ["10:00 start log [correlationId: TC-001-abc]"]})

    write_evidence_tabs(str(path), [ev], "2026-06-04 21:00:00")

    assert "TC-001" in load_workbook(path).sheetnames
    text = _tab_text(path, "TC-001")
    assert "RESULT: ✅ PASS" in text
    assert "2026-06-04 21:00:00" in text
    assert "https://api.test/orders" in text
    assert "start log [correlationId: TC-001-abc]" in text
    # definitions still parse to the two original cases (evidence tab ignored by the parser)
    assert {c.test_id for c in read_test_suite(str(path)).cases} == {"TC-001", "TC-002"}


def test_evidence_tab_overwrites_not_appends(tmp_path):
    path = tmp_path / "suite.xlsx"
    _make_suite_xlsx(path)

    write_evidence_tabs(str(path), [_evidence(201, {"x": ["first run line"]})], "2026-06-04 21:00")
    write_evidence_tabs(str(path), [_evidence(500, {"x": ["second run line"]})], "2026-06-04 22:00")

    # exactly one TC-001 tab, reflecting only the latest run
    assert load_workbook(path).sheetnames.count("TC-001") == 1
    text = _tab_text(path, "TC-001")
    assert "second run line" in text
    assert "first run line" not in text
    assert "2026-06-04 22:00" in text


def test_evidence_blank_when_no_log_lines_match(tmp_path):
    path = tmp_path / "suite.xlsx"
    _make_suite_xlsx(path)

    write_evidence_tabs(str(path), [_evidence(201, {"never matched": []})], "2026-06-04 21:00")

    ws = load_workbook(path)["TC-001"]
    rows = [["" if c is None else str(c) for c in r] for r in ws.iter_rows(values_only=True)]
    needle_rows = [r for r in rows if r and r[0] == "never matched"]
    assert needle_rows and needle_rows[0][1] == ""  # the expected string is listed, evidence blank


def test_safe_sheet_name_sanitizes_and_dedupes():
    used: set[str] = set()
    assert _safe_sheet_name("TC-001", used) == "TC-001"
    assert _safe_sheet_name("a/b:c*d", used) == "a_b_c_d"  # forbidden chars replaced
    long_id = "X" * 40
    assert len(_safe_sheet_name(long_id, used)) <= 31  # truncated to Excel's limit
    first = _safe_sheet_name("dup", used)
    second = _safe_sheet_name("dup", used)
    assert first != second  # collisions disambiguated


def test_write_evidence_tabs_noop_for_numbers(tmp_path):
    # non-xlsx suffix is a silent no-op (summary block handles .numbers separately)
    path = tmp_path / "suite.numbers"
    path.write_bytes(b"not a real numbers file")
    write_evidence_tabs(str(path), [_evidence(201, {"x": []})], "2026-06-04 21:00")
    assert path.read_bytes() == b"not a real numbers file"  # untouched
