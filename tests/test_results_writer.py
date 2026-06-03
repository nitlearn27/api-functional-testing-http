"""results_writer: appends a timestamped block and never corrupts the definitions."""

from openpyxl import Workbook, load_workbook

from api_log_test_mcp.models import CaseReport, SuiteReport
from api_log_test_mcp.tools.results_writer import ResultsWriteError, write_results
from api_log_test_mcp.tools.suite import read_test_suite


def _make_suite_xlsx(path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Basepath", "https://api.test/"])
    ws.append([])
    ws.append(["test_id", "method", "url", "expected_status"])
    ws.append(["TC-001", "POST", "/orders", 201])
    ws.append(["TC-002", "POST", "/orders", 400])
    wb.save(path)


def _report() -> SuiteReport:
    return SuiteReport(
        total=2, passed=1, failed=1,
        cases=[
            CaseReport(test_id="TC-001", passed=True, actual_status=201, expected_status=201,
                       correlation_id="TC-001-evid01"),
            CaseReport(test_id="TC-002", passed=False, actual_status=201, expected_status=400,
                       correlation_id="TC-002-evid02"),
        ],
    )


def test_write_results_appends_block_and_preserves_cases(tmp_path):
    path = tmp_path / "suite.xlsx"
    _make_suite_xlsx(path)

    run_at = write_results(str(path), _report(), run_at="2026-06-03 21:00:00")
    assert run_at == "2026-06-03 21:00:00"

    ws = load_workbook(path).active
    rows = [[c if c is not None else "" for c in r] for r in ws.iter_rows(values_only=True)]
    flat = ["|".join(str(c) for c in r) for r in rows]

    assert any("RESULTS — run 2026-06-03 21:00:00" in line for line in flat)
    assert any(line.startswith("TC-001|PASS|201|201|TC-001-evid01") for line in flat)
    assert any(line.startswith("TC-002|FAIL|201|400|TC-002-evid02") for line in flat)
    assert any("correlation_id" in line for line in flat)  # header column present

    # the suite still parses to exactly the two original cases (results block ignored)
    suite = read_test_suite(str(path))
    assert {c.test_id for c in suite.cases} == {"TC-001", "TC-002"}


def test_write_results_is_idempotent_in_structure(tmp_path):
    """Two runs append two blocks; the case section is unchanged either time."""
    path = tmp_path / "suite.xlsx"
    _make_suite_xlsx(path)
    write_results(str(path), _report(), run_at="2026-06-03 21:00:00")
    write_results(str(path), _report(), run_at="2026-06-03 21:30:00")

    ws = load_workbook(path).active
    flat = ["|".join(str(c) for c in r if c is not None) for r in ws.iter_rows(values_only=True)]
    assert sum("RESULTS — run" in line for line in flat) == 2
    suite = read_test_suite(str(path))
    assert {c.test_id for c in suite.cases} == {"TC-001", "TC-002"}


def test_write_results_restores_on_corruption(tmp_path, monkeypatch):
    """If the save mangles the definition rows, the file is restored and an error raised."""
    path = tmp_path / "suite.xlsx"
    _make_suite_xlsx(path)
    original = load_workbook(path).active
    original_rows = [tuple(r) for r in original.iter_rows(values_only=True)]

    # Simulate a corrupting writer: it appends the block but also wrecks a definition cell.
    import api_log_test_mcp.tools.results_writer as rw

    real_append = rw._append_xlsx

    def corrupting_append(file_path, block):
        real_append(file_path, block)
        wb = load_workbook(file_path)
        ws = wb.active
        ws["A4"] = "MANGLED"  # clobber TC-001's test_id
        wb.save(file_path)

    monkeypatch.setattr(rw, "_append_xlsx", corrupting_append)

    try:
        write_results(str(path), _report())
        raised = False
    except ResultsWriteError:
        raised = True

    assert raised
    # file restored to original definition rows
    restored_rows = [tuple(r) for r in load_workbook(path).active.iter_rows(values_only=True)]
    assert restored_rows == original_rows
