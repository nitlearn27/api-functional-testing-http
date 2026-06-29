"""run_suite end-to-end with the HTTP layer mocked (no network, no logs)."""

import functools
import json
from pathlib import Path

import httpx
from openpyxl import Workbook

from api_log_test_mcp.tools import orchestrate
from api_log_test_mcp.tools.http_runner import call_api as real_call_api

HEADERS = [
    "test_id", "method", "url", "headers", "body", "auth_required",
    "expected_status", "expected_response", "response_match_mode", "validate_logs",
]

ROWS = [
    ["TC-001", "POST", "/orders", "", json.dumps({"sku": "ABC-100", "qty": 2}), "no",
     201, json.dumps({"status": "ACCEPTED", "sku": "ABC-100"}), "json_subset", "no"],
    ["TC-002", "POST", "/orders", "", json.dumps({"sku": "ABC-100"}), "no",
     400, json.dumps({"error": "VALIDATION_ERROR", "field": "qty"}), "json_subset", "no"],
]


def _make_suite(tmp_path: Path) -> str:
    wb = Workbook()
    ws = wb.active
    ws.append(["Basepath", "https://api.test/"])
    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)
    path = tmp_path / "suite.xlsx"
    wb.save(path)
    return str(path)


def _mock_handler(request: httpx.Request) -> httpx.Response:
    payload = json.loads(request.content)
    if "qty" in payload:
        return httpx.Response(201, json={"status": "ACCEPTED", "sku": payload["sku"]})
    return httpx.Response(400, json={"error": "VALIDATION_ERROR", "field": "qty"})


def test_run_suite_all_pass(tmp_path, monkeypatch):
    client = httpx.Client(transport=httpx.MockTransport(_mock_handler))
    monkeypatch.setattr(
        orchestrate, "call_api", functools.partial(real_call_api, client=client)
    )

    report = orchestrate.run_suite(_make_suite(tmp_path))

    assert report.total == 2
    assert report.passed == 2
    assert report.failed == 0
    assert all(c.log_validation is None for c in report.cases)  # validate_logs=no -> skipped


def test_run_suite_reports_failure(tmp_path, monkeypatch):
    # Always return 500 so both cases fail their status assertion.
    def fail_handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(500, json={"error": "boom"})

    client = httpx.Client(transport=httpx.MockTransport(fail_handler))
    monkeypatch.setattr(
        orchestrate, "call_api", functools.partial(real_call_api, client=client)
    )

    report = orchestrate.run_suite(_make_suite(tmp_path))
    assert report.passed == 0
    assert report.failed == 2
    assert all(not c.passed for c in report.cases)


def test_run_and_record_writes_separate_results_file(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    client = httpx.Client(transport=httpx.MockTransport(_mock_handler))
    monkeypatch.setattr(
        orchestrate, "call_api", functools.partial(real_call_api, client=client)
    )

    suite_path = _make_suite(tmp_path)
    suite_bytes_before = Path(suite_path).read_bytes()

    report, run_at, results_path = orchestrate.run_and_record(suite_path)

    assert report.total == 2 and report.passed == 2
    # Results land in a SEPARATE <stem>_results.xlsx, never the suite.
    assert results_path == str(tmp_path / "suite_results.xlsx")
    assert Path(results_path).exists()
    # The suite file itself is left byte-for-byte unchanged.
    assert Path(suite_path).read_bytes() == suite_bytes_before
    # The results file carries the case definitions plus a RESULTS block and per-case tabs.
    wb = load_workbook(results_path)
    data_ws = wb.worksheets[0]
    firsts = [str(r[0]) for r in data_ws.iter_rows(values_only=True) if r and r[0] is not None]
    assert "TC-001" in firsts
    assert any(f.startswith("RESULTS") for f in firsts)
    assert "TC-001" in wb.sheetnames  # evidence tab
