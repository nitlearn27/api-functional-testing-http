"""create_test_suite_from_schema: generic OpenAPI walk -> .xlsx, round-tripping the parser.

Offline (pure file I/O). Uses a self-contained spec (no external fixture) that exercises the
generic per-operation coverage — query params, **header params**, path params, body rules, errors
and auth — and asserts the generated sheet parses with zero errors.
"""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml

from api_log_test_mcp.tools.suite import read_test_suite
from api_log_test_mcp.tools.suite_generator import generate_test_suite

SPEC: dict = {
    "openapi": "3.0.0",
    "info": {"title": "Items", "version": "1.0"},
    "servers": [{"url": "https://api.example.com/v1"}],
    "security": [{"bearerAuth": []}],
    "components": {
        "securitySchemes": {"bearerAuth": {"type": "http", "scheme": "bearer"}},
        "schemas": {
            "ErrorResponse": {
                "type": "object",
                "required": ["code", "message"],
                "properties": {"code": {"type": "string"}, "message": {"type": "string"}},
            },
            "Item": {
                "type": "object",
                "required": ["id", "name"],
                "properties": {"id": {"type": "string"}, "name": {"type": "string"}},
            },
        },
    },
    "paths": {
        "/items": {
            "get": {
                "summary": "List items",
                "parameters": [
                    {"name": "q", "in": "query", "required": True, "schema": {"type": "string"}},
                    {
                        "name": "kind",
                        "in": "query",
                        "schema": {"type": "string", "enum": ["a", "b"]},
                    },
                    {
                        "name": "X-Tenant",
                        "in": "header",
                        "required": True,
                        "schema": {"type": "string"},
                    },
                ],
                "responses": {
                    "200": {
                        "description": "ok",
                        "content": {
                            "application/json": {
                                "schema": {
                                    "type": "array",
                                    "items": {"$ref": "#/components/schemas/Item"},
                                }
                            }
                        },
                    },
                    "400": {
                        "description": "bad",
                        "content": {
                            "application/json": {
                                "schema": {"$ref": "#/components/schemas/ErrorResponse"}
                            }
                        },
                    },
                },
            },
            "post": {
                "summary": "Create item",
                "requestBody": {
                    "required": True,
                    "content": {
                        "application/json": {
                            "schema": {
                                "type": "object",
                                "required": ["name"],
                                "properties": {
                                    "name": {"type": "string", "pattern": "^[A-Z]"},
                                    "status": {"type": "string", "enum": ["new", "used"]},
                                    "tags": {
                                        "type": "array",
                                        "maxItems": 2,
                                        "items": {"type": "string"},
                                    },
                                },
                            }
                        }
                    },
                },
                "responses": {
                    "201": {
                        "description": "created",
                        "content": {
                            "application/json": {"schema": {"$ref": "#/components/schemas/Item"}}
                        },
                    },
                    "400": {
                        "description": "bad",
                        "content": {
                            "application/json": {
                                "schema": {"$ref": "#/components/schemas/ErrorResponse"}
                            }
                        },
                    },
                },
            },
        },
        "/items/{id}": {
            "get": {
                "summary": "Get item",
                "parameters": [
                    {
                        "name": "id",
                        "in": "path",
                        "required": True,
                        "schema": {"type": "string", "pattern": "^ITM-"},
                    }
                ],
                "responses": {
                    "200": {
                        "description": "ok",
                        "content": {
                            "application/json": {"schema": {"$ref": "#/components/schemas/Item"}}
                        },
                    },
                    "404": {
                        "description": "missing",
                        "content": {
                            "application/json": {
                                "schema": {"$ref": "#/components/schemas/ErrorResponse"}
                            }
                        },
                    },
                },
            }
        },
    },
}


@pytest.fixture
def generated(tmp_path: Path) -> tuple[dict, str]:
    spec_path = tmp_path / "items.yaml"
    spec_path.write_text(yaml.safe_dump(SPEC))
    out = tmp_path / "items_suite.xlsx"
    summary = generate_test_suite(str(spec_path), str(out))
    return summary, str(out)


def test_writes_sheet_with_spec_basepath(generated):
    summary, out = generated
    assert Path(out).exists()
    assert summary["base_path"] == "https://api.example.com/v1"  # straight from servers[0].url
    assert summary["case_count"] >= 10  # comprehensive coverage across 3 operations


def test_worksheet_is_named_tests(generated):
    from openpyxl import load_workbook

    _, out = generated
    assert load_workbook(out).sheetnames == ["tests"]


def test_round_trips_through_parser_without_errors(generated):
    summary, out = generated
    suite = read_test_suite(out)
    assert suite.base_path == summary["base_path"]
    assert suite.parse_errors == []
    assert len(suite.cases) == summary["case_count"]


def test_covers_every_validation_category(generated):
    _, out = generated
    suite = read_test_suite(out)
    statuses = {c.expected_status for c in suite.cases}
    assert {200, 201, 400, 401, 404, 415} <= statuses
    assert 422 not in statuses  # body-validation folded into 400 (Mulesoft cannot return 422)

    descriptions = " || ".join(c.description or "" for c in suite.cases)
    assert "missing required query 'q'" in descriptions  # required query-param omission
    assert "missing required header 'X-Tenant'" in descriptions  # required header omission (new)
    assert "not in allowed enum" in descriptions
    assert "violates pattern" in descriptions
    assert "exceeds maxItems" in descriptions


def test_positive_sends_required_query_param_and_header(generated):
    _, out = generated
    suite = read_test_suite(out)
    pos = next(
        c for c in suite.cases if c.method == "GET" and c.expected_status == 200 and "?" in c.url
    )
    assert "q=" in pos.url
    assert "X-Tenant" in pos.headers


def test_error_cases_use_the_spec_error_envelope(generated):
    _, out = generated
    suite = read_test_suite(out)
    errors = [c for c in suite.cases if c.expected_status in {400, 401, 404, 415}]
    assert errors
    for case in errors:
        assert set(case.expected_response) == {"code", "message"}, case.test_id


def test_all_body_validation_maps_to_400(generated):
    _, out = generated
    suite = read_test_suite(out)
    body_negs = [
        c
        for c in suite.cases
        if c.method == "POST"
        and any(
            k in (c.description or "")
            for k in ("missing required", "violates pattern", "not in allowed enum", "maxItems")
        )
    ]
    assert body_negs
    assert all(c.expected_status == 400 for c in body_negs)


def test_malformed_body_case_kept_as_raw_string(generated):
    _, out = generated
    suite = read_test_suite(out)
    malformed = [c for c in suite.cases if "malformed JSON" in (c.description or "")]
    assert malformed
    assert isinstance(malformed[0].body, str)
